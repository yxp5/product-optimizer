# A software project that optimize product finding with respect to customer's requirements
# Author: yxp5, Liu Jing, Wang Sibo
# Supervisor: Tang Chaojuan

import os
import encode
import random
import openpyxl
import tkinter as tk
from tkinter import filedialog
from meta import *

def ExcelToInt(string):
    """
    str -> int
    
    Given Excel column format string, return the equivalent integer value.
    
    Ex: "A" -> 1, "AA" -> 27, "Z" -> 26
    """
    total = 0
    for pos, char in enumerate(string[::-1]):
        if ord(char) < ord('A') or ord(char) > ord('Z'): raise ValueError("Wrong input! The string is not of Excel column format!")
        total += (ord(char)-ord('A')+1) * (26 ** pos)
    return total

def IntToExcel(integer):
    """
    int -> string
    
    Given integer value, return the equivalent Excel column format string.
    
    Ex: 1 -> "A", 27 -> "AA", 26 -> "Z"
    """
    if integer < 1: raise ValueError("Wrong input! The integer must be at least 1!")
    string = ""
    while integer > 0:
        integer -= 1
        remainder = integer % 26
        string = chr(ord('A') + remainder) + string
        integer = integer // 26
    return string

def ReadExcel(filename):
    """
    Read data from an Excel file [filename], not called explicitly
    """
    global FEATURES, FEATMAP, PRODUCTS, PRODMAP, HP
    data = openpyxl.load_workbook(filename)
    df = data.active

    matrix = []
    for i in range(df.max_row):
        matrix.append([])
        for j,col in enumerate(df.iter_cols(1, df.max_column)):
            matrix[i].append(col[i].value)

    # Extract features
    ftmp = []
    fmatrix = matrix[HP["ROW"]-1:HP["END"]]
    
    tmp = []
    for row in fmatrix:
        if row[HP["SUM"]-1] != "SUM":
            fname = row[HP["FNAME"]-1] if HP["FNAME"] != -1 else "None"
            propname = row[HP["PROPNAME"]-1] if HP["PROPNAME"] != -1 else "None"
            desname = row[HP["DES"]-1] if HP["DES"] != -1 else "None"
            tmp.append([fname, propname, desname])
    fmatrix = tmp

    for f in fmatrix:
        prop = IsUnique(f[1])
        feat = Feature(f[0].replace('\n', ''), prop, f[2])
        ftmp.append(feat)
        FEATMAP.update({feat.uid: feat})

    # Extract products and add features
    prtmp = []
    pmatrix = matrix[HP["PNAME"]-1]
    pmatrix = pmatrix[HP["COL"]-1:]

    dmatrix = matrix[HP["ROW"]-1:HP["END"]]
    dmatrix = [row[HP["COL"]-1:] for row in dmatrix if row[0] != "SUM"]

    # Extract for each column/product, the amount of row/feature
    for j, p in enumerate(pmatrix):
        req = {}
        for i, f in enumerate(ftmp):
            cell = dmatrix[i][j] if dmatrix[i][j] else 0
            cell = cell if type(cell) == int else eval(f"{cell[1:]}")
            req.update({f.uid: cell})
        product = Product(p.replace('\n', ''), req)
        prtmp.append(product)
        PRODMAP.update({product.uid: product})
    
    # Update the data only when no error occured to avoid information loss
    FEATURES = ftmp
    PRODUCTS = prtmp
    Response(f"Excel file [{filename}] successfully loaded")
    return UpdateScreen()

def Save():
    """
    Save current data file to a software-bound data file type ".popt" (maybe replace by a PostgreS database later)
    for faster future load than reading an Excel file
    """
    global FEATURES, PRODUCTS, PROPERTIES, DEFAULT_FILE
    
    def subSave():
        filename = save_entry.get()
        # Save to filename
        if len(filename) < 5 or filename[-5:] != SUFFIX:
            raise NameError(f"Wrong file format to save to. Destination [{filename}] must end with [{SUFFIX}]")
        
        fp = open(filename, "w")
        
        tmp = []
        for p in PROPERTIES:
            tmp.append(f"{p.name},{p.uid}")
        string = "\n".join(tmp)
        
        string += "\n;\n"
            
        tmp = []
        for f in FEATURES:
            tmp.append(f"{f.name},{f.prop},{f.desc},{f.uid}")
        string += "\n".join(tmp)
        
        string += "\n;\n"
        
        tmp = []
        for p in PRODUCTS:
            s = []
            for fid, need in p.req.items():
                s.append(f"{fid}:{need}")
            req = ".".join(s)
            tmp.append(f"{p.name},{req},{p.uid}")
        string += "\n".join(tmp)
        
        fp.write(string)
        
        Response(f"Data file [{filename}] successfully saved")
        DEFAULT_FILE = filename
        return subExit()
    
    def subSaveSelect():
        current_directory = os.getcwd()
        path = filedialog.askopenfilename(parent=save_window, initialdir=current_directory, title='Please select a file')
        save_entry.delete(0, tk.END)
        save_entry.insert(0, path)
        return
    
    def subExit():
        return save_window.destroy()
    
    # Let client input save file path
    save_window = tk.Toplevel()
    save_window.title("Saving data")
    save_window.resizable(False, False)
    
    save_frame = tk.Frame(save_window, bg="gray", width=500)
    
    save_entry = tk.Entry(save_frame, width=50, bg="white", fg="black")
    save_button = tk.Button(save_frame, text="Save", command=subSave)
    save_select_button = tk.Button(save_frame, text="Select", command=subSaveSelect)
    exit_button = tk.Button(save_frame, text="Exit", command=subExit)
    save_entry.grid(row=0, column=2, sticky="nwe", padx=5, pady=5)
    save_button.grid(row=0, column=0, sticky="w", padx=5, pady=5)
    save_select_button.grid(row=0, column=1, sticky="w", padx=5, pady=5)
    exit_button.grid(row=0, column=3, sticky="w", padx=5, pady=5)
    
    save_frame.pack(fill=tk.BOTH, side=tk.LEFT)
    save_window.mainloop()

def Load(quick=False):
    """
    Load data from either ".xlsx" or ".popt" file suffix
    """
    global FEATURES, PRODUCTS, PROPERTIES, FEATMAP, PRODMAP, PROPMAP, DEFAULT_FILE
    # Check for quick load mode, if quick load, use the default file "data.popt"
    if quick and DEFAULT_FILE:
        filename = DEFAULT_FILE
    else:
        current_directory = os.getcwd()
        filename = filedialog.askopenfilename(parent=main_window, initialdir=current_directory, title='Please select a file')
    
    # If file is Excel file, then switch to ReadExcel
    if filename[-5:] == ".xlsx": return ReadExcel(filename)
    
    fp = open(filename, "r")
    data = fp.read()
    fp.close()
    
    ftmp, prtmp = [], []
    pd, fd, prd = data.split("\n;\n")
    
    # Load properties
    for prop in pd.split("\n"):
        pname, pid = prop.split(",")
        IsUnique(pname)
    
    # Load features
    for feat in fd.split("\n"):
        fname, fprop, fdesc, fid = feat.split(",")
        feature = Feature(fname, IsUnique(fprop), fdesc)
        ftmp.append(feature)
    
    # Load products
    for prod in prd.split("\n"):
        prname, prreq, prid = prod.split(",")
        req = {}
        for entry in prreq.split("."):
            fid, need = entry.split(":")
            req.update({fid: int(need)})
        product = Product(prname, req)
        prtmp.append(product)
    
    # Update the data only when no error occured to avoid information loss
    FEATURES = ftmp
    PRODUCTS = prtmp
    
    # Then update mapping
    for f in ftmp:
        FEATMAP.update({f.uid: f})
    for pr in prtmp:
        PRODMAP.update({pr.uid: pr})
    
    Response(f"Data file [{filename}] successfully loaded")
    DEFAULT_FILE = filename # Set current data file to be [filename]
    return UpdateScreen() # Update according to best products

def QuickLoad():
    return Load(True)

def Order(subpage=0):
    """
    Big function to edit client order
    """
    global ORDER
    if not ORDER: raise ValueError("Load some data first!")
    
    # WHY CAN'T I ACCESS VARIABLE WITH MORE THAN 1 LAYER DIFFERENCE!?! PYTHON SCOPE MAN...
    class SP:
        def __init__(self, subpage):
            self.value = subpage
    sp = SP(subpage)
    
    buttons = []
    entries = []
    
    def subInit():
        # Button formatting
        limit = 15
        low = sp.value * 20
        high = min(low + 20, len(FEATURES))
        for index in range(low, high):
            feature_button = tk.Button(features_frame, anchor="w", width=96, text="Untitled", command=None)
            feature_button.config(font=("Menlo", 14), fg="black")
            
            count_entry = tk.Entry(features_frame, bg="white", width=8)
            count_entry.config(font=("Menlo", 14), fg="black")
            count_entry.insert(0, -1)
            
            feature_button.grid(row=index, column=0, sticky="nsew", padx=2, pady=2)
            count_entry.grid(row=index, column=1, sticky="nsew", padx=2, pady=2)
            
            buttons.append(feature_button)
            entries.append(count_entry)
        
        return subUpdateScreen()
    
    def subUpdateScreen():
        # Widget updating
        limit = 15
        low = sp.value * 20
        high = min(low + 20, len(FEATURES))
        for index in range(low, high):
            feature_button = buttons[index-low]
            count_entry = entries[index-low]
            
            feature = FEATURES[index]
            trimmed_title = f"{index+1:<4} {feature.name:<20s} {feature.prop.name:^35s} {feature.uid:^30s}"
            feature_button.config(text=trimmed_title)
            
            count_entry.delete(0, tk.END)
            count_entry.insert(0, ORDER[feature.uid])
            
            feature_button.grid(row=index, column=0, sticky="nsew", padx=2, pady=2)
            count_entry.grid(row=index, column=1, sticky="nsew", padx=2, pady=2)
        
        return
    
    def subPrevious():
        subSaveNeed()
        if sp.value == 0: return
        
        sp.value -= 1
        return subUpdateScreen()
    
    def subNext():
        subSaveNeed()
        if (sp.value+1) * 20 >= len(FEATURES): return

        sp.value += 1
        return subUpdateScreen()

    def subSaveNeed(exit_toggle=False):
        low = sp.value * 20
        high = min(low + 20, len(FEATURES))
        
        # Getting modified feature count
        needs = []
        for widget in features_frame.winfo_children():
            if isinstance(widget, tk.Entry):
                if not widget.get().isnumeric():
                    return Response("Please input only integers for your order")
                
                needs.append(int(widget.get()))
        
        # Save
        for index in range(low, high):
            ORDER[FEATURES[index].uid] = needs[index-low]
        
        if exit_toggle: subExit()
        return
    
    def subExit(print_toggle=True):
        subSaveNeed()
        
        if print_toggle: Response("Exit editing order")
        UpdateScreen()
        return order_window.destroy()
    
    order_window = tk.Toplevel()
    order_window.title("Edit quizlet")
    order_window.geometry("900x700")
    order_window.resizable(False, False)
    order_window.config(bg="black")
    order_window.protocol("WM_DELETE_WINDOW", lambda exit_toggle=True: subSaveNeed(exit_toggle))
    
    features_frame = tk.Frame(order_window, bg="white", height=640, width=890)
    setting_frame = tk.Frame(order_window, bg="gray", height=40, width=890)
    features_frame.grid_propagate(False)
    
    subInit()
    
    # Setting frame
    prev_button = tk.Button(setting_frame, text="Prev Page", command=subPrevious)
    next_button = tk.Button(setting_frame, text="Next Page", command=subNext)
    save_button = tk.Button(setting_frame, text="Save", command=subSaveNeed)
    exit_button = tk.Button(setting_frame, text="Exit", command=subExit)
    
    prev_button.grid(row=0, column=0, padx=5, pady=5)
    next_button.grid(row=0, column=1, padx=5, pady=5)
    save_button.grid(row=0, column=2, padx=5, pady=5)
    exit_button.grid(row=0, column=3, padx=5, pady=5)
    
    features_frame.grid(row=0, column=0, sticky="nswe", padx=5, pady=5)
    setting_frame.grid(row=1, columnspan=2, sticky="nswe", padx=5, pady=5)
    
    order_window.mainloop()
    return

def Previous():
    """
    Go to the previous page of products
    """
    global PAGE
    if PAGE == 0: return
    
    PAGE -= 1
    return UpdateScreen()

def Next():
    """
    Go to the next page of products
    """
    global PAGE
    if (PAGE+1) * 33 >= len(RESULTS[0]): return
    
    PAGE += 1
    return UpdateScreen()

def Setting():
    global HP
    
    def subInit():
        for index, (hp, value) in enumerate(HP.items()):
            hp_button = tk.Button(hp_frame, anchor="w", text=hp, command=None)
            hp_button.config(font=("Menlo", 14), fg="black")
            
            value_entry = tk.Entry(hp_frame, bg="white", width=5)
            value_entry.config(font=("Menlo", 14), fg="black")
            value_entry.insert(0, value)
            
            hp_button.grid(row=index, column=0, sticky="nsew", padx=2, pady=2)
            value_entry.grid(row=index, column=1, sticky="nsew", padx=2, pady=2)
        
        return
    
    def subSaveSetting(exit_toggle=False):
        global HP
        
        values = []
        for widget in hp_frame.winfo_children():
            if isinstance(widget, tk.Entry):
                try:
                    values.append(int(widget.get()))
                except ValueError:
                    return Response("Please input only integers as hyperparameters")
        
        # Save
        for index, key in enumerate(HP.keys()):
            HP[key] = values[index]
        
        if exit_toggle: subExit()
        return Response("Hyperparameters updated")
    
    def subExit():
        return setting_window.destroy()
    
    setting_window = tk.Toplevel()
    setting_window.title("Setting")
    setting_window.geometry("200x320")
    setting_window.resizable(False, False)
    setting_window.config(bg="black")
    setting_window.protocol("WM_DELETE_WINDOW", lambda exit_toggle=True: subSaveSetting(exit_toggle))
    
    hp_frame = tk.Frame(setting_window, bg="white", height=260, width=200)
    setting_frame = tk.Frame(setting_window, bg="gray", height=50, width=200)
    hp_frame.grid_propagate(False)
    
    subInit()
    
    # Setting frame
    save_button = tk.Button(setting_frame, text="Save", command=subSaveSetting)
    exit_button = tk.Button(setting_frame, text="Exit", command=subExit)
    
    save_button.grid(row=0, column=0, padx=5, pady=5)
    exit_button.grid(row=0, column=1, padx=5, pady=5)
    
    hp_frame.grid(row=0, column=0, sticky="nswe", padx=5, pady=5)
    setting_frame.grid(row=1, columnspan=2, sticky="nswe", padx=5, pady=5)
    
    setting_window.mainloop()
    return

def Exit():
    """
    Destroy main_window and exit on code 1
    """
    main_window.destroy()
    exit(1)

# Helper functions
def IsUnique(name):
    """
    Given a property name [name], generate a Property class object [prop] and check its uniqueness with respect to set [PSET]
    If [prop.uid] is new, then register it in PSET and return [prop]
    Otherwise, return the existing property with same uid as [prop] using [PROPMAP]
    """
    global PSET, PROPERTIES, PROPMAP
    prop = Property(str(name).replace('\n', ''))
    
    # Register if the property is new
    if prop.uid not in PSET:
        PSET.add(prop.uid)
        PROPERTIES.append(prop)
        PROPMAP.update({prop.uid: prop})
        return prop
    # Else return the already registered property
    else:
        return PROPMAP[prop.uid]

def PrintProduct(p):
    """
    Return a detailed information string [info] about Product class object [p]
    """
    global FEATMAP
    info = f"Product name:\n\t{p.name}\n\nRequirements:\n"
        
    for feat, need in p.req.items():
        if need > 0: info += f"{FEATMAP[feat].Info(True)}\n\tNeed: {need}\n\n"
        
    return info

def Response(text):
    """
    Append the string [text] in the response Text widget [resp_text]
    """
    global resp_text
    
    resp_text.config(state="normal")
    resp_text.insert(tk.END, f"\n{text}")
    resp_text.config(state="disabled")
    
    return

def ShowProduct(product):
    """
    Big function that opens a new window on click to show all the information about Product class object [product]
    """
    def subUpdateScreen():
        info_text.config(state="normal")
        info_text.delete("1.0", tk.END)
        info_text.insert("1.0", PrintProduct(product))
        info_text.config(state="disabled")
        return
    
    def subExit(print_toggle=True):
        if print_toggle: Response(f"Exit viewing product {product.name}")
        UpdateScreen()
        return product_window.destroy()
    
    product_window = tk.Toplevel()
    product_window.title("Product Viewing")
    product_window.geometry("660x500")
    product_window.resizable(False, False)
    product_window.config(bg="black")
    
    features_frame = tk.Frame(product_window, bg="white", height=640, width=890)
    setting_frame = tk.Frame(product_window, bg="gray", height=40, width=890)
    features_frame.grid_propagate(False)
    
    info_text = tk.Text(features_frame, bg="light gray")
    info_text.pack(fill=tk.BOTH, side=tk.LEFT, expand=True)
    info_text.config(font=("Times New Roman", 16), fg="black", state="disabled")
    
    subUpdateScreen()
    
    # Setting frame
    exit_button = tk.Button(setting_frame, text="Exit", command=subExit)
    exit_button.grid(row=0, column=0, padx=5, pady=5)
    
    features_frame.grid(row=0, column=0, sticky="nswe", padx=5, pady=5)
    setting_frame.grid(row=1, columnspan=2, sticky="nswe", padx=5, pady=5)
    
    product_window.mainloop()
    return

def UpdateScreen():
    """
    Display a portion of all products in the output frame [outp_frame]
    """
    global RESULTS, outp_frame
    for button in outp_frame.winfo_children(): button.destroy()
    
    if not ORDER: InitOrder()
    BestProducts(ORDER)
    
    best, rest = RESULTS
    limit = 25
    per_row = 3
    per_col = 11
    low = PAGE * per_row * per_col
    high = min(low + 33, len(best))
    for index in range(low, high):
        product = best[index]
        trimmed = f"{index+1}) {product.name}"
        
        if len(trimmed) <= limit:
            padding = " " * (limit - len(trimmed))
            trimmed = trimmed + padding
        else:
            trimmed = trimmed[:limit-3] + "..."
        
        product_button = tk.Button(outp_frame, text=f"{trimmed}",
                                    command=lambda product=product: ShowProduct(product))
        product_button.config(font=("Menlo", 14), fg="black")
        product_button.grid(row=index%per_col, column=index//per_col, sticky="w", padx=8, pady=8)
    return

def InitOrder():
    """
    Initialize default client order to be 0 for all available features
    """
    global ORDER
    
    ORDER = {}
    for f in FEATURES:
        ORDER.update({f.uid: 0})
    
    return

def BestProducts(order):
    global PRODUCTS, FEATMAP, RESULTS

    best = []
    rest = {}
    for p in PRODUCTS:
        good = True
        diff = {}
        for fid, need in order.items():
            if p.req[fid] < need:
                diff.update({FEATMAP[fid].name: need - p.req[fid]})
                good = False
        if good: best.append(p)
        else: rest.update({p.name: diff})
    
    RESULTS = (best, rest)
    return

# Functions for testing purpose (not adapted in tkinter
def GenerateRandomOrder():
    global FEATURES
    
    order = {}
    randfeats = random.sample(FEATURES, random.randint(5,10))
    randneeds = [random.randint(1,10) for f in randfeats]
    
    for feat, need in zip(randfeats, randneeds):
        order.update({feat.uid: need})
    
    return order

def Test(order=None):
    if not order: order = GenerateRandomOrder()
    results, rest = BestProducts(order)
    
    print(f"Customer order: {order}\nBest products:\n")
    if len(results):
        for p in results:
            print(p.name)
    else:
        print("None")
    
    print(f"Insatisfactory products' difference:\n")
    if len(rest):
        for pname, diff in rest.items():
            print(f"Product: {pname}")
            for fname, lack in diff.items():
                print(f"\t{fname}: {lack}")
    else:
        print("None")
    
    return

# Global variables
FILENAME = "data.xlsx" # For faster testing purpose
FEATURES = []
PRODUCTS = []
RESULTS = ([], [])
PROPERTIES = []
PSET = set()
FEATMAP = {}
PRODMAP = {}
PROPMAP = {}
DEFAULT_FILE = "data.popt"
SUFFIX = ".popt"
PAGE = 0
ORDER = None

# Definable hyper parameters, number are inuitive natural numbers, not computer array index number -> conversion +/- 1
HP = {"COL": 4, "ROW": 3, "SUM": 1, "DES": -1, "PNAME": 1, "FNAME": 2, "PROPNAME": 3, "END": 151}
# COL, ROW: data starting position
# SUM: SUM column number
# DES: Description column number
# PNAME: Product name row
# FNAME: Feature name col
# PROPNAME: Property name col, usually current intensity (Ample)
# END: End row of required data

# Initializing software structure
# Main window
main_window = tk.Tk()
main_window.title("UAES XE/ESF2 Product Optimizer")
main_window.geometry("900x700")
main_window.resizable(False, False)

main_window.rowconfigure(0, minsize=600, weight=1)
main_window.columnconfigure(1, minsize=800, weight=1)

# Frames inside main window
main_frame = tk.Frame(main_window, bg="black")
butt_frame = tk.Frame(main_window, relief=tk.RAISED, bd=2, bg="gray")
main_frame.grid_propagate(False)

# outp_title for spacing (not important), outp_frame for product list
outp_title = tk.Frame(main_frame, bg="black", height=5)
outp_frame = tk.Frame(main_frame, bg="white", height=500)

# io_title for spacing (not important), io_frame for logger/print
io_title = tk.Frame(main_frame, bg="black", height=5)
io_frame = tk.Frame(main_frame, bg="light gray", height=300)

# Essential buttons
btn_sav = tk.Button(butt_frame, text="Save", command=Save)
btn_loa = tk.Button(butt_frame, text="Load", command=Load)
btn_qld = tk.Button(butt_frame, text="Quick Load", command=QuickLoad)
btn_run = tk.Button(butt_frame, text="Order", command=Order)
btn_prv = tk.Button(butt_frame, text="Prev Page", command=Previous)
btn_nxt = tk.Button(butt_frame, text="Next Page", command=Next)
btn_set = tk.Button(butt_frame, text="Setting", command=Setting)
btn_ext = tk.Button(butt_frame, text="Exit", command=Exit)

btn_sav.grid(row=1, column=0, sticky="ew", padx=5, pady=5)
btn_loa.grid(row=2, column=0, sticky="ew", padx=5, pady=5)
btn_qld.grid(row=3, column=0, sticky="ew", padx=5, pady=5)
btn_run.grid(row=4, column=0, sticky="ew", padx=5, pady=5)
btn_prv.grid(row=5, column=0, sticky="ew", padx=5, pady=5)
btn_nxt.grid(row=6, column=0, sticky="ew", padx=5, pady=5)
btn_set.grid(row=7, column=0, sticky="ew", padx=5, pady=5)
btn_ext.grid(row=8, column=0, sticky="ew", padx=5, pady=5)

butt_frame.grid(row=0, column=0, sticky="ns")
main_frame.grid(row=0, column=1, sticky="nswe")
outp_title.pack(fill=tk.BOTH)
outp_frame.pack(fill=tk.BOTH, padx=5, expand=True)
outp_frame.grid_propagate(False)
io_title.pack(fill=tk.BOTH)
io_frame.pack(fill=tk.BOTH, padx=5, expand=True)

# Initialize logger/print text widget
resp_text = tk.Text(io_frame, width=50, bg="light gray")
resp_text.pack(fill=tk.BOTH, side=tk.LEFT, expand=True)
resp_text.config(font=("Times New Roman", 16), fg="black")

# Welcome!
resp_text.insert("1.0", "A software project by yxp5, Liu Jing, Wang Sibo under supervision of Tang Chaojuan :)")
resp_text.config(state="disabled")

main_window.mainloop()













